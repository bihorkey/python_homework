import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ========================
# 1. 数据预处理
# ========================
def preprocess_data(file_path):
    """加载并清洗数据，合并细分行业为大类"""
    try:
        df = pd.read_csv(file_path)
    except FileNotFoundError:
        print(f"错误：未找到数据文件 '{file_path}'")
        return None
    except Exception as e:
        print(f"数据读取失败：{str(e)}")
        return None

    # 处理财富值为数值类型
    df['财富值_人民币_亿'] = pd.to_numeric(df['财富值_人民币_亿'], errors='coerce')

    # 行业合并规则（覆盖200+细分行业，合并为12大领域）
    industry_mapping = {
        '房地产': ['房地产', '地产', '置业', '物业', '不动产', '房产'],
        '投资': ['投资', '金融投资', '资本', '资产管理', '私募', '风投', '创投'],
        '医药': ['医药', '生物制药', '医疗器械', '医疗服务', '制药', '生物科技', '医疗健康'],
        '互联网': ['互联网', '电子商务', '网络服务', '社交媒体', '游戏', '云计算', '大数据', '人工智能', '在线'],
        '制造业': ['制造业', '制造', '工业', '机械设备', '装备制造', '工程机械'],
        '新能源': ['新能源', '锂电池', '光伏', '太阳能', '风能', '储能', '电池', '新能源汽车'],
        '半导体': ['半导体', '芯片', '集成电路', '微电子'],
        '消费': ['饮料', '食品', '餐饮', '酒类', '零售', '百货', '商超', '快消品', '消费品', '食品饮料'],
        '金融': ['金融服务', '银行', '保险', '证券', '信托', '支付', '金融科技'],
        '教育': ['教育', '培训', '在线教育'],
        '物流': ['物流', '快递', '运输', '供应链'],
        '化工': ['化工', '化学', '材料', '新材料', '石化'],
        '电子': ['电子', '消费电子', '智能硬件', '家电', '电子产品'],
        '汽车': ['汽车', '汽车制造', '汽车零部件', '整车'],
        '农业': ['农业', '农产品', '养殖', '种植', '渔业'],
        '文化娱乐': ['文化', '娱乐', '影视', '传媒', '体育', '音乐', '游戏'],
        '建筑': ['建筑', '建材', '装饰', '工程', '施工'],
        '能源': ['能源', '电力', '煤炭', '石油', '天然气'],
        '环保': ['环保', '环境', '水处理', '废物处理'],
        '其他': []
    }

    def map_industry(industry):
        """将细分行业映射到大类"""
        if pd.isna(industry):
            return '其他'
        industry = str(industry).strip()
        for main_cat, sub_cats in industry_mapping.items():
            if any(sub in industry for sub in sub_cats):
                return main_cat
        return '其他'

    df['行业分类'] = df['所在行业_中文'].apply(map_industry)

    # 过滤无效数据
    cleaned_df = df.dropna(subset=['财富值_人民币_亿', '行业分类'])
    print(f"数据预处理完成：原始数据 {len(df)} 条 → 有效数据 {len(cleaned_df)} 条")
    return cleaned_df


# ========================
# 2. 行业统计分析
# ========================
def analyze_industries(df):
    """计算行业核心指标：人数、总财富、平均财富、财富集中度等"""
    # 分组统计基础指标
    industry_stats = df.groupby('行业分类').agg(
        富豪人数=('行业分类', 'count'),
        总财富_亿=('财富值_人民币_亿', 'sum'),
        平均财富_亿=('财富值_人民币_亿', 'mean'),
        财富中位数_亿=('财富值_人民币_亿', 'median'),
        最大财富_亿=('财富值_人民币_亿', 'max')
    ).reset_index()

    # 计算「财富集中度」：行业前3名富豪财富占比
    def top3_share(group):
        if len(group) == 0:
            return 0
        top3_sum = group.nlargest(3, keep='first').sum()
        total_sum = group.sum()
        return top3_sum / total_sum if total_sum > 0 else 0

    top3_shares = df.groupby('行业分类', group_keys=False)['财富值_人民币_亿'].apply(top3_share).reset_index()
    top3_shares.columns = ['行业分类', '财富集中度']
    industry_stats = pd.merge(industry_stats, top3_shares, on='行业分类', how='left')

    # 提取除“其他”行业外的有效行业数据
    valid_industries = industry_stats[industry_stats['行业分类'] != '其他'].copy()

    if not valid_industries.empty:
        # 计算有效行业的占比（富豪数、财富）
        total_people = valid_industries['富豪人数'].sum()
        total_wealth = valid_industries['总财富_亿'].sum()
        valid_industries['富豪占比'] = valid_industries['富豪人数'] / total_people
        valid_industries['财富占比'] = valid_industries['总财富_亿'] / total_wealth

        # 按总财富对有效行业进行排序
        valid_industries = valid_industries.sort_values('总财富_亿', ascending=False).reset_index(drop=True)
        valid_industries.index += 1  # 序号从1开始

        # 处理“其他”行业数据
        other_industry = industry_stats[industry_stats['行业分类'] == '其他'].copy()
        if not other_industry.empty:
            # 计算“其他”行业占比，分母包含“其他”行业
            other_industry['富豪占比'] = other_industry['富豪人数'] / (
                        total_people + other_industry['富豪人数'].values[0])
            other_industry['财富占比'] = other_industry['总财富_亿'] / (
                        total_wealth + other_industry['总财富_亿'].values[0])
            # 将“其他”行业放在最后，索引为有效行业数量 + 1
            other_industry.index = [len(valid_industries) + 1]
            industry_stats = pd.concat([valid_industries, other_industry])
        else:
            industry_stats = valid_industries
    else:
        # 如果没有有效行业，直接返回原统计结果（可能只有“其他”行业）
        industry_stats['富豪占比'] = 1.0
        industry_stats['财富占比'] = 1.0
        industry_stats.index += 1

    return industry_stats


# ========================
# 3. 生成行业发展洞察
# ========================
def generate_insights(industry_stats):
    """从统计结果推导行业发展态势"""
    insights = []
    if len(industry_stats) == 0:
        return ["⚠️ 无有效行业数据，无法分析"]

    # 排除"其他"行业进行洞察分析
    valid_industries = industry_stats[industry_stats['行业分类'] != '其他']
    if valid_industries.empty:
        return ["⚠️ 有效行业数据不足，无法进行洞察分析"]

    # 1. 主导行业分析
    top_industry = valid_industries.iloc[0]
    insights.append(
        f"1. 主导行业：{top_industry['行业分类']}以{top_industry['总财富_亿']:.0f}亿人民币总财富领跑，"
        f"占有效行业总财富的{top_industry['财富占比'] * 100:.1f}%，聚集{top_industry['富豪人数']}位富豪"
    )

    # 2. 高价值领域分析（人均财富前25%）
    try:
        high_threshold = valid_industries['平均财富_亿'].quantile(0.75)
        high_growth = valid_industries[valid_industries['平均财富_亿'] > high_threshold]
        if not high_growth.empty:
            industries = ', '.join(high_growth['行业分类'].tolist())
            median_avg = high_growth['平均财富_亿'].median()
            insights.append(
                f"2. 高价值领域：{industries}等行业人均财富超{median_avg:.1f}亿，呈现资本密集型特征"
            )
        else:
            insights.append("2. 高价值领域：各行业财富分布均衡，未现显著高价值赛道")
    except Exception as e:
        insights.append(f"2. 高价值领域：分析失败（{str(e)}）")

    # 3. 头部效应分析（财富集中度前3）
    try:
        concentrated = valid_industries.nlargest(3, '财富集中度')
        if not concentrated.empty:
            top_concent = concentrated.iloc[0]
            insights.append(
                f"3. 头部效应：{top_concent['行业分类']}行业财富集中度达{top_concent['财富集中度'] * 100:.1f}%，头部企业垄断性强"
            )
        else:
            insights.append("3. 头部效应：各行业财富分散，头部效应不明显")
    except Exception as e:
        insights.append(f"3. 头部效应：分析失败（{str(e)}）")

    # 4. 潜力赛道分析（多强竞争：富豪数>10，集中度<30%，人均>中位数）
    try:
        median_avg = valid_industries['平均财富_亿'].median()
        emerging = valid_industries[
            (valid_industries['富豪人数'] > 10) &
            (valid_industries['财富集中度'] < 0.3) &
            (valid_industries['平均财富_亿'] > median_avg)
            ]
        if not emerging.empty:
            industries = ', '.join(emerging['行业分类'].tolist())
            insights.append(
                f"4. 潜力赛道：{industries}等行业呈现多强竞争格局，财富分布均衡，成长空间广阔"
            )
        else:
            insights.append("4. 潜力赛道：未发现明显均衡发展的潜力行业")
    except Exception as e:
        insights.append(f"4. 潜力赛道：分析失败（{str(e)}）")

    # 5. 其他行业说明
    other_industry = industry_stats[industry_stats['行业分类'] == '其他']
    if not other_industry.empty:
        other = other_industry.iloc[0]
        insights.append(
            f"5. 未分类领域：未明确归类的行业总财富为{other['总财富_亿']:.0f}亿人民币，"
            f"占整体的{other['财富占比'] * 100:.1f}%，涉及{other['富豪人数']}位富豪"
        )

    return insights


# ========================
# 5. 导出Excel报告
# ========================
def export_to_excel(industry_stats, insights, output_path='2024胡润百富榜行业分析报告.xlsx'):
    """将分析结果导出到Excel文件，包含格式化的表格和关键发现"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "行业分析报告"

    # 设置表头样式
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=16)
    border_style = Side(style='thin', color='FF000000')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 添加标题
    ws['A1'] = "2024胡润百富榜行业分析报告"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')

    # 添加分析样本信息
    ws['A2'] = f"分析样本：共{industry_stats['富豪人数'].sum()}位富豪，覆盖{len(industry_stats)}个行业类别"
    ws['A2'].font = Font(size=12)
    ws.merge_cells('A2:H2')

    # 空行分隔
    ws.append([])

    # 定义列标题和格式
    columns = [
        ('序号', 6, 'int'),
        ('行业分类', 15, 'text'),
        ('富豪人数', 12, 'int'),
        ('总财富_亿', 15, 'float'),
        ('平均财富_亿', 15, 'float'),
        ('财富集中度', 15, 'percent'),
        ('富豪占比', 12, 'percent'),
        ('财富占比', 12, 'percent')
    ]

    # 添加表头
    header_row = ws.append([col[0] for col in columns])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = Border(
            left=border_style,
            right=border_style,
            top=border_style,
            bottom=border_style
        )

    # 添加数据行
    for idx, row in industry_stats.iterrows():
        data_row = [
            idx,
            row['行业分类'],
            row['富豪人数'],
            row['总财富_亿'],
            row['平均财富_亿'],
            row['财富集中度'],
            row['富豪占比'],
            row['财富占比']
        ]
        ws.append(data_row)

        # 设置单元格格式和边框
        for col_idx, (_, _, col_type) in enumerate(columns, 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = Border(
                left=border_style,
                right=border_style,
                top=border_style,
                bottom=border_style
            )

            # 设置单元格格式
            if col_type == 'int':
                cell.number_format = '0'
            elif col_type == 'float':
                cell.number_format = '0.00'
            elif col_type == 'percent':
                cell.number_format = '0.00%'

    # 设置列宽
    for col_idx, (_, width, _) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 空行分隔
    ws.append([])
    ws.append([])

    # 添加关键发现标题
    ws.append(["关键发现："])
    ws[ws.max_row][0].font = title_font
    ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')

    # 添加关键发现内容
    for insight in insights:
        ws.append([insight])
        ws[ws.max_row][0].font = Font(size=12)
        ws.merge_cells(f'A{ws.max_row}:H{ws.max_row}')

    # 保存Excel文件
    wb.save(output_path)
    print(f"\n✅ Excel报告已保存至：{output_path}")


# ========================
# 4. 主程序
# ========================
if __name__ == "__main__":
    # 数据文件路径（需替换为实际路径）
    DATA_PATH = '2024胡润百富榜.csv'
    # 输出Excel路径
    OUTPUT_PATH = '2024胡润百富榜行业分析报告.xlsx'

    # 数据预处理
    df_clean = preprocess_data(DATA_PATH)
    if df_clean is None:
        exit(1)

    # 行业统计
    industry_stats = analyze_industries(df_clean)

    # 打印行业排名表格
    print("\n" + "=" * 70)
    print("2024胡润百富榜行业分析报告")
    print("=" * 70)
    print(f"分析样本：共{len(df_clean)}位富豪，覆盖{len(industry_stats)}个行业类别\n")

    # 表格格式配置，根据实际列内容长度合理调整宽度
    column_widths = {
        '序号': 6,
        '行业分类': 12,
        '富豪人数': 10,
        '总财富_亿': 12,
        '平均财富_亿': 12,
        '财富集中度': 12,
        '富豪占比': 10,
        '财富占比': 10
    }
    # 构建表头
    header_format = (
        "{:<{序号_width}}"
        "{:<{行业分类_width}}"
        "{:<{富豪人数_width}}"
        "{:<{总财富_亿_width}}"
        "{:<{平均财富_亿_width}}"
        "{:<{财富集中度_width}}"
        "{:<{富豪占比_width}}"
        "{:<{财富占比_width}}"
    ).format(
        '序号', '行业分类', '富豪人数', '总财富_亿', '平均财富_亿', '财富集中度', '富豪占比', '财富占比',
        序号_width=column_widths['序号'],
        行业分类_width=column_widths['行业分类'],
        富豪人数_width=column_widths['富豪人数'],
        总财富_亿_width=column_widths['总财富_亿'],
        平均财富_亿_width=column_widths['平均财富_亿'],
        财富集中度_width=column_widths['财富集中度'],
        富豪占比_width=column_widths['富豪占比'],
        财富占比_width=column_widths['财富占比']
    )
    print(header_format)
    # 构建分隔线
    separator = '-' * sum(column_widths.values())
    print(separator)

    # 逐行格式化输出表格内容
    for idx, row in industry_stats.iterrows():
        row_format = (
            "{:<{序号_width}}"
            "{:<{行业分类_width}}"
            "{:<{富豪人数_width}}"
            "{:<{总财富_亿_width}.1f}"
            "{:<{平均财富_亿_width}.2f}"
            "{:<{财富集中度_width}.2%}"
            "{:<{富豪占比_width}.2%}"
            "{:<{财富占比_width}.2%}"
        ).format(
            idx,
            row['行业分类'][:column_widths['行业分类']],  # 截断过长行业名
            row['富豪人数'],
            row['总财富_亿'],
            row['平均财富_亿'],
            row['财富集中度'],
            row['富豪占比'],
            row['财富占比'],
            序号_width=column_widths['序号'],
            行业分类_width=column_widths['行业分类'],
            富豪人数_width=column_widths['富豪人数'],
            总财富_亿_width=column_widths['总财富_亿'],
            平均财富_亿_width=column_widths['平均财富_亿'],
            财富集中度_width=column_widths['财富集中度'],
            富豪占比_width=column_widths['富豪占比'],
            财富占比_width=column_widths['财富占比']
        )
        print(row_format)

    # 打印关键发现
    insights = generate_insights(industry_stats)
    print("\n关键发现：")
    for i, insight in enumerate(insights, 1):
        print(f"{i}. {insight}")

    # 导出Excel报告
    export_to_excel(industry_stats, insights, OUTPUT_PATH)